from __future__ import annotations

import base64
from dataclasses import dataclass
from datetime import datetime, timedelta
import hashlib
from io import BytesIO
from pathlib import Path
import random
import string

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


st.set_page_config(page_title="Excel Practice Bot", page_icon=":bar_chart:", layout="wide", initial_sidebar_state="collapsed")
DEFAULT_LOGO_PATH = Path(__file__).parent / "assets" / "excel_logics_logo.png"
ALT_LOGO_PATH = Path(__file__).parent / "assets" / "Logo1.png"
QUESTION_REGISTRY_PATH = Path(__file__).parent / "data" / "used_question_signatures.txt"
PARTICIPANT_SEQUENCE_PATH = Path(__file__).parent / "data" / "participant_sequence.txt"


@dataclass
class QuestionSpec:
    question_id: str
    question: str
    answer: str


STANDARD_QUESTION_COUNT = 20


def _random_phone_ok(phone: str) -> bool:
    clean = "".join(ch for ch in phone if ch.isdigit())
    return 10 <= len(clean) <= 15


def _question_signature(question_text: str) -> str:
    normalized = " ".join(question_text.lower().split())
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def _load_used_question_signatures() -> set[str]:
    if not QUESTION_REGISTRY_PATH.exists():
        return set()
    return {line.strip() for line in QUESTION_REGISTRY_PATH.read_text(encoding="utf-8").splitlines() if line.strip()}


def _persist_used_question_signatures(signatures: set[str]) -> None:
    QUESTION_REGISTRY_PATH.parent.mkdir(parents=True, exist_ok=True)
    existing = _load_used_question_signatures()
    merged = existing | signatures
    QUESTION_REGISTRY_PATH.write_text("\n".join(sorted(merged)) + "\n", encoding="utf-8")


def _next_queue_id() -> str:
    PARTICIPANT_SEQUENCE_PATH.parent.mkdir(parents=True, exist_ok=True)
    current = 0
    if PARTICIPANT_SEQUENCE_PATH.exists():
        raw = PARTICIPANT_SEQUENCE_PATH.read_text(encoding="utf-8").strip()
        if raw.isdigit():
            current = int(raw)
    next_num = current + 1
    PARTICIPANT_SEQUENCE_PATH.write_text(str(next_num), encoding="utf-8")
    return f"ELST-{next_num:03d}"


def _resolve_serial_number(profile: dict[str, str]) -> str:
    existing = profile.get("SerialNumber") or profile.get("QueueID")
    if existing and str(existing).strip():
        return str(existing).strip()
    # Backfill for older workbooks so certificate never shows N/A.
    new_serial = _next_queue_id()
    profile["SerialNumber"] = new_serial
    return new_serial


def _hex_to_rgb(color_hex: str) -> tuple[int, int, int]:
    color_hex = color_hex.lstrip("#")
    return tuple(int(color_hex[i : i + 2], 16) for i in (0, 2, 4))


def _resolve_logo_bytes(logo_upload) -> tuple[bytes | None, str | None]:
    if logo_upload is not None:
        return logo_upload.getvalue(), logo_upload.type or "image/png"
    for logo_path in [DEFAULT_LOGO_PATH, ALT_LOGO_PATH]:
        if logo_path.exists():
            suffix = logo_path.suffix.lower()
            if suffix in [".jpg", ".jpeg"]:
                mime = "image/jpeg"
            else:
                mime = "image/png"
            return logo_path.read_bytes(), mime
    return None, None


def _logo_data_uri(logo_bytes: bytes | None, mime: str | None) -> str:
    if not logo_bytes or not mime:
        return ""
    b64 = base64.b64encode(logo_bytes).decode("utf-8")
    return f"data:{mime};base64,{b64}"


def _placeholder_logo() -> Image.Image:
    img = Image.new("RGBA", (240, 80), (255, 255, 255, 0))
    d = ImageDraw.Draw(img)
    d.rounded_rectangle((4, 8, 72, 72), radius=12, fill=(179, 205, 54, 255))
    d.rounded_rectangle((30, 8, 72, 72), radius=12, fill=(0, 153, 204, 255))
    return img


def _pick_font(size: int) -> ImageFont.ImageFont:
    candidates = [
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Helvetica.ttc",
    ]
    for path in candidates:
        if Path(path).exists():
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
    return ImageFont.load_default()


def _build_scorecard_png(
    brand_name: str,
    logo_bytes: bytes | None,
    serial_number: str,
    student_name: str,
    practice_role: str,
    percentage: float,
    correct_count: int,
    incorrect_count: int,
    primary_color: str,
    secondary_color: str,
    exam_name: str,
    result_status: str,
) -> bytes:
    w, h = 1400, 900
    img = Image.new("RGB", (w, h), "white")
    draw = ImageDraw.Draw(img)
    primary_rgb = _hex_to_rgb(primary_color)
    secondary_rgb = _hex_to_rgb(secondary_color)

    draw.rounded_rectangle((30, 30, w - 30, h - 30), radius=28, fill=(248, 250, 252), outline=primary_rgb, width=4)
    draw.rectangle((30, 30, w - 30, 185), fill=(255, 255, 255))
    draw.rectangle((30, 185, w - 30, 205), fill=secondary_rgb)
    draw.line((30, 185, w - 30, 185), fill=primary_rgb, width=4)

    title_font = _pick_font(56)
    sub_font = _pick_font(34)
    body_font = _pick_font(30)
    big_font = _pick_font(92)

    if logo_bytes:
        try:
            logo_img = Image.open(BytesIO(logo_bytes)).convert("RGBA")
            logo_img.thumbnail((220, 120))
            img.paste(logo_img, (60, 52), logo_img)
        except Exception:
            pass

    draw.text((300, 60), f"{brand_name}", font=title_font, fill=primary_rgb)
    draw.text((300, 122), "Excel Practice Score Card", font=sub_font, fill=(30, 41, 59))

    draw.text((90, 260), "Issued To:", font=sub_font, fill=(30, 41, 59))
    draw.text((90, 312), student_name or "Student", font=title_font, fill=primary_rgb)
    draw.text((90, 366), f"Serial Number: {serial_number}", font=body_font, fill=(15, 23, 42))

    draw.text((90, 420), f"Exam: {exam_name}", font=body_font, fill=(15, 23, 42))
    draw.text((90, 468), f"Practice Role: {practice_role}", font=body_font, fill=(15, 23, 42))
    draw.text((90, 516), f"Result: {result_status}", font=body_font, fill=(15, 23, 42))
    draw.text((90, 564), f"Correct: {correct_count}    Incorrect: {incorrect_count}", font=body_font, fill=(15, 23, 42))
    draw.text((90, 612), f"Issued On: {datetime.now().strftime('%Y-%m-%d')}", font=body_font, fill=(15, 23, 42))

    draw.rounded_rectangle((930, 290, 1300, 650), radius=26, fill=(255, 255, 255), outline=secondary_rgb, width=4)
    draw.text((1020, 350), "Score", font=sub_font, fill=primary_rgb)
    draw.text((980, 430), f"{percentage:.2f}%", font=big_font, fill=(17, 24, 39))

    draw.text((90, 730), f"Certified by {brand_name}", font=sub_font, fill=primary_rgb)
    draw.text((90, 778), "Upload this score card on LinkedIn and tag Excel Logics", font=body_font, fill=(51, 65, 85))

    out = BytesIO()
    img.save(out, format="PNG")
    out.seek(0)
    return out.getvalue()


def _generate_base_frame(role: str, rows: int, seed: int) -> pd.DataFrame:
    random.seed(seed)
    np.random.seed(seed)

    start_date = datetime.today() - timedelta(days=180)
    dates = [start_date + timedelta(days=int(x)) for x in np.random.randint(0, 180, rows)]

    regions = ["North", "South", "East", "West", "Central"]
    departments = ["Ops", "Finance", "Sales", "HR", "Support"]
    categories = ["A", "B", "C", "D"]
    products = ["Laptop", "Mouse", "Keyboard", "Monitor", "Dock"]
    processes = ["Onboarding", "Support", "Dispatch", "Returns", "Audit"]
    topics = ["Formulas", "Pivot", "Lookup", "Charts", "Cleaning"]
    levels = ["Beginner", "Intermediate", "Advanced"]
    expense_types = ["Travel", "Utilities", "Software", "Maintenance", "Training"]

    df = pd.DataFrame(
        {
            "RecordID": [f"R{100000 + i}" for i in range(rows)],
            "Date": pd.to_datetime(dates),
            "Region": np.random.choice(regions, rows),
        }
    )

    if role == "Data Analyst":
        df["Department"] = np.random.choice(departments, rows)
        df["Category"] = np.random.choice(categories, rows)
        df["Quantity"] = np.random.randint(1, 50, rows)
        df["UnitPrice"] = np.round(np.random.uniform(10, 500, rows), 2)
        df["Discount"] = np.round(np.random.uniform(0, 0.2, rows), 3)
        df["Revenue"] = np.round(df["Quantity"] * df["UnitPrice"] * (1 - df["Discount"]), 2)
    elif role == "Sales Executive":
        df["SalesRep"] = np.random.choice(["Anita", "Rahul", "Kiran", "Nisha", "Vivek"], rows)
        df["Product"] = np.random.choice(products, rows)
        df["Quantity"] = np.random.randint(1, 40, rows)
        df["UnitPrice"] = np.round(np.random.uniform(20, 800, rows), 2)
        df["Discount"] = np.round(np.random.uniform(0, 0.25, rows), 3)
        df["Revenue"] = np.round(df["Quantity"] * df["UnitPrice"] * (1 - df["Discount"]), 2)
    elif role == "HR Associate":
        df["Department"] = np.random.choice(departments, rows)
        df["ExperienceYears"] = np.round(np.random.uniform(0.5, 15, rows), 1)
        df["CTC"] = np.round(np.random.uniform(3_00_000, 24_00_000, rows), 0)
        df["BonusPercent"] = np.round(np.random.uniform(0.02, 0.25, rows), 3)
        df["AnnualBonus"] = np.round(df["CTC"] * df["BonusPercent"], 0)
    elif role == "Finance Associate":
        df["CostCenter"] = np.random.choice(["CC-101", "CC-104", "CC-202", "CC-305"], rows)
        df["ExpenseType"] = np.random.choice(expense_types, rows)
        df["Amount"] = np.round(np.random.uniform(500, 50000, rows), 2)
        df["TaxPercent"] = np.round(np.random.uniform(0.05, 0.18, rows), 3)
        df["GrossAmount"] = np.round(df["Amount"] * (1 + df["TaxPercent"]), 2)
    elif role == "Operations Associate":
        df["Site"] = np.random.choice(["Plant-A", "Plant-B", "WH-01", "WH-02"], rows)
        df["Process"] = np.random.choice(processes, rows)
        df["TicketsHandled"] = np.random.randint(5, 200, rows)
        df["SLAHours"] = np.round(np.random.uniform(1, 48, rows), 1)
        df["DelayHours"] = np.round(np.maximum(df["SLAHours"] - np.random.uniform(0, 30, rows), 0), 1)
    else:
        df["Topic"] = np.random.choice(topics, rows)
        df["Level"] = np.random.choice(levels, rows)
        df["Score"] = np.random.randint(35, 100, rows)
        df["PracticeHours"] = np.round(np.random.uniform(0.5, 10, rows), 1)

    return df.sort_values("Date").reset_index(drop=True)


def _format_ans(val: float | int | str) -> str:
    if isinstance(val, str):
        return val
    if isinstance(val, (int, np.integer)):
        return str(int(val))
    return f"{float(val):.2f}"


def _build_question_bank(df: pd.DataFrame, seed: int, used_signatures: set[str]) -> tuple[list[QuestionSpec], set[str]]:
    rng = random.Random(seed)
    qa_pairs: list[tuple[str, str]] = []
    qa_pairs.append(("How many records are present in PracticeData?", _format_ans(len(df))))
    qa_pairs.append(("How many unique values are present in Region?", _format_ans(df["Region"].nunique())))
    qa_pairs.append(("What is the total count of records in Region = North?", _format_ans(int((df["Region"] == "North").sum()))))
    qa_pairs.append(("What is the earliest date in Date column?", str(df["Date"].min().date())))
    qa_pairs.append(("What is the latest date in Date column?", str(df["Date"].max().date())))

    by_region_counts = df.groupby("Region", as_index=False).size().sort_values("size", ascending=False)
    qa_pairs.append(("Which Region has the highest number of records?", str(by_region_counts.iloc[0]["Region"])))
    qa_pairs.append(("What is the count of records in the top Region?", _format_ans(int(by_region_counts.iloc[0]["size"]))))

    numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    for col in numeric_cols:
        qa_pairs.append((f"What is SUM of `{col}`?", _format_ans(df[col].sum())))
        qa_pairs.append((f"What is AVERAGE of `{col}`?", _format_ans(df[col].mean())))
        qa_pairs.append((f"What is MIN of `{col}`?", _format_ans(df[col].min())))
        qa_pairs.append((f"What is MAX of `{col}`?", _format_ans(df[col].max())))
        qa_pairs.append((f"What is MEDIAN of `{col}`?", _format_ans(df[col].median())))
        qa_pairs.append((f"What is the 75th percentile (P75) of `{col}`?", _format_ans(df[col].quantile(0.75))))
        for region in ["North", "South", "East", "West", "Central"]:
            regional = df.loc[df["Region"] == region, col]
            if len(regional) > 0:
                qa_pairs.append((f"What is SUM of `{col}` for Region = {region}?", _format_ans(regional.sum())))
                qa_pairs.append((f"What is AVERAGE of `{col}` for Region = {region}?", _format_ans(regional.mean())))

    text_cols = [c for c in df.columns if c not in ["RecordID", "Date"] and df[c].dtype == "object"]
    for col in text_cols:
        top = df[col].value_counts().idxmax()
        qa_pairs.append((f"Which value appears most in `{col}`?", str(top)))
        qa_pairs.append((f"How many unique values are there in `{col}`?", _format_ans(df[col].nunique())))

    # Build a large dynamic pool so questions can stay unique for a high number of participants.
    for col in numeric_cols:
        col_min = float(df[col].min())
        col_max = float(df[col].max())
        if col_min == col_max:
            continue
        for _ in range(25):
            threshold = round(rng.uniform(col_min, col_max), 3)
            qa_pairs.append(
                (f"For `{col}`, how many rows have value greater than {threshold}?", _format_ans(int((df[col] > threshold).sum())))
            )
            qa_pairs.append(
                (f"For `{col}`, how many rows have value less than {threshold}?", _format_ans(int((df[col] < threshold).sum())))
            )
            qa_pairs.append(
                (
                    f"For `{col}`, what is SUM where value is greater than {threshold}?",
                    _format_ans(float(df.loc[df[col] > threshold, col].sum())),
                )
            )

    if "Date" in df.columns:
        date_min = df["Date"].min()
        date_max = df["Date"].max()
        day_span = max(int((date_max - date_min).days), 1)
        for _ in range(30):
            offset_a = rng.randint(0, day_span)
            offset_b = rng.randint(offset_a, day_span)
            start = (date_min + timedelta(days=offset_a)).date()
            end = (date_min + timedelta(days=offset_b)).date()
            count = int(((df["Date"].dt.date >= start) & (df["Date"].dt.date <= end)).sum())
            qa_pairs.append((f"How many records have Date between {start} and {end} (inclusive)?", _format_ans(count)))

    rng.shuffle(qa_pairs)

    selected_pairs: list[tuple[str, str]] = []
    selected_signatures: set[str] = set()
    for question_text, answer in qa_pairs:
        sig = _question_signature(question_text)
        if sig in used_signatures or sig in selected_signatures:
            continue
        selected_pairs.append((question_text, answer))
        selected_signatures.add(sig)
        if len(selected_pairs) == STANDARD_QUESTION_COUNT:
            break

    # Fallback generator for extra uniqueness.
    attempts = 0
    while len(selected_pairs) < STANDARD_QUESTION_COUNT and attempts < 500:
        attempts += 1
        if numeric_cols:
            col = rng.choice(numeric_cols)
            col_min = float(df[col].min())
            col_max = float(df[col].max())
            threshold = round(rng.uniform(col_min, col_max), 5)
            op = rng.choice([">", "<", ">=", "<="])
            if op == ">":
                mask = df[col] > threshold
            elif op == "<":
                mask = df[col] < threshold
            elif op == ">=":
                mask = df[col] >= threshold
            else:
                mask = df[col] <= threshold
            question_text = f"For `{col}`, how many rows have value {op} {threshold}?"
            answer = _format_ans(int(mask.sum()))
        else:
            suffix = f"{rng.randint(0, 999999):06d}"
            question_text = f"How many RecordID values end with `{suffix}`?"
            answer = _format_ans(int(df["RecordID"].astype(str).str.endswith(suffix).sum()))

        sig = _question_signature(question_text)
        if sig in used_signatures or sig in selected_signatures:
            continue
        selected_pairs.append((question_text, answer))
        selected_signatures.add(sig)

    questions = [QuestionSpec(str(i + 1), q, a) for i, (q, a) in enumerate(selected_pairs)]
    return questions, selected_signatures


def _student_seed(student_name: str, phone: str, target_role: str) -> int:
    nonce = datetime.now().isoformat(timespec="microseconds")
    raw = f"{student_name.strip().lower()}|{phone.strip()}|{target_role}|{nonce}"
    digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    # NumPy legacy seed must be within [0, 2**32 - 1]
    return int(digest[:12], 16) % (2**32 - 1)


def _profile_frame(payload: dict[str, str]) -> pd.DataFrame:
    return pd.DataFrame({"Field": list(payload.keys()), "Value": [str(v) for v in payload.values()]})


def _build_workbook_bytes(profile: dict[str, str], df: pd.DataFrame, questions: list[QuestionSpec]) -> bytes:
    questions_df = pd.DataFrame(
        [{"QuestionID": q.question_id, "Question": q.question, "ExpectedAnswer": q.answer} for q in questions]
    )
    student_q_df = questions_df[["QuestionID", "Question"]].copy()
    answer_sheet_df = questions_df[["QuestionID"]].copy()
    answer_sheet_df["YourAnswer"] = ""
    key_df = questions_df[["QuestionID", "ExpectedAnswer"]].copy()

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _profile_frame(profile).to_excel(writer, index=False, sheet_name="StudentProfile")
        df.to_excel(writer, index=False, sheet_name="PracticeData")
        student_q_df.to_excel(writer, index=False, sheet_name="Questions")
        answer_sheet_df.to_excel(writer, index=False, sheet_name="Answers")
        key_df.to_excel(writer, index=False, sheet_name="__KEY")

        wb = writer.book
        wb["__KEY"].sheet_state = "hidden"

    output.seek(0)
    return output.getvalue()


def _to_float_safe(value: str) -> float | None:
    try:
        return float(str(value).strip())
    except Exception:
        return None


def _evaluate_workbook(upload_bytes: bytes) -> tuple[pd.DataFrame, float, int, int, dict[str, str]]:
    wb = load_workbook(BytesIO(upload_bytes), data_only=True)
    if "Answers" not in wb.sheetnames or "__KEY" not in wb.sheetnames:
        raise ValueError("Required sheets `Answers` and `__KEY` are missing.")

    answers_ws = wb["Answers"]
    key_ws = wb["__KEY"]

    answers = {}
    for row in answers_ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        answers[str(row[0]).strip()] = "" if row[1] is None else str(row[1]).strip()

    key = {}
    for row in key_ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        key[str(row[0]).strip()] = "" if row[1] is None else str(row[1]).strip()

    profile: dict[str, str] = {}
    if "StudentProfile" in wb.sheetnames:
        profile_ws = wb["StudentProfile"]
        for row in profile_ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            profile[str(row[0]).strip()] = "" if row[1] is None else str(row[1]).strip()

    expected_total = int(profile.get("QuestionCount", str(STANDARD_QUESTION_COUNT)))
    if len(key) != expected_total:
        raise ValueError(
            f"This workbook has {len(key)} questions, but expected {expected_total}. "
            "Please upload the latest generated workbook."
        )

    rows = []
    score = 0
    for qid, expected in key.items():
        got = answers.get(qid, "")
        status = "Incorrect"

        exp_num = _to_float_safe(expected)
        got_num = _to_float_safe(got)
        if exp_num is not None and got_num is not None:
            if abs(exp_num - got_num) <= 0.01:
                status = "Correct"
        elif expected.strip().lower() == got.strip().lower():
            status = "Correct"

        if status == "Correct":
            score += 1

        rows.append(
            {
                "QuestionID": qid,
                "ExpectedAnswer": expected,
                "StudentAnswer": got,
                "Result": status,
            }
        )

    report_df = pd.DataFrame(rows)
    report_df["QuestionNo"] = pd.to_numeric(report_df["QuestionID"], errors="coerce")
    report_df = report_df.sort_values("QuestionNo").reset_index(drop=True)
    total = len(report_df) if len(report_df) else 1
    incorrect = total - score
    percent = round((score / total) * 100, 2)
    return report_df, percent, score, incorrect, profile


def _inject_brand_css(primary_color: str, secondary_color: str) -> None:
    st.markdown(
        f"""
        <style>
        :root {{
            --brand-primary: {primary_color};
            --brand-secondary: {secondary_color};
            --bg-main: #f8fafc;
            --bg-card: #ffffff;
            --text-main: #0f172a;
            --text-soft: #334155;
            --border-soft: #dbe4ee;
        }}
        .stApp {{
            background:
                radial-gradient(circle at 12% 8%, rgba(171, 199, 56, 0.18), transparent 34%),
                radial-gradient(circle at 92% 5%, rgba(23, 152, 197, 0.16), transparent 32%),
                var(--bg-main);
            color: var(--text-main);
        }}
        h1, h2, h3, h4, h5, h6, p, label {{
            color: var(--text-main);
        }}
        [data-testid="stHeader"] {{
            background: #ffffff;
            border-bottom: 1px solid var(--border-soft);
        }}
        [data-testid="stSidebar"] {{
            display: none !important;
        }}
        [data-testid="stSidebarCollapsedControl"] {{
            display: none !important;
        }}
        [data-testid="stToolbar"] {{
            display: none !important;
        }}
        [data-testid="collapsedControl"] {{
            display: none !important;
        }}
        [data-testid="stTextInputRootElement"] input,
        [data-testid="stTextArea"] textarea,
        [data-testid="stSelectbox"] div[data-baseweb="select"] > div,
        [data-testid="stNumberInput"] input {{
            background: #ffffff !important;
            color: var(--text-main) !important;
            border: 1px solid #cbd5e1 !important;
        }}
        [data-testid="stFileUploader"] section {{
            background: #ffffff;
            border: 1px dashed #94a3b8;
        }}
        .certificate-card {{
            background: linear-gradient(120deg, #ffffff, #f0f9ff);
            border: 1px solid #bfd9e7;
            border-radius: 16px;
            padding: 20px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.08);
        }}
        .certificate-title {{
            margin: 0;
            font-size: 1.2rem;
            color: var(--brand-primary);
            font-weight: 700;
        }}
        .certificate-score {{
            margin: 6px 0 0 0;
            font-size: 2rem;
            font-weight: 800;
            color: var(--text-main);
        }}
        .certificate-sub {{
            margin: 4px 0;
            color: var(--text-soft);
        }}
        .cert-logo {{
            width: 110px;
            height: auto;
            display: block;
            margin-bottom: 8px;
            background: #ffffff;
            padding: 6px;
            border-radius: 10px;
            border: 1px solid #dbe4ee;
        }}
        .stDownloadButton > button, .stButton > button {{
            border-radius: 10px;
            border: 1px solid #0f172a !important;
            color: #ffffff !important;
            background: #0f172a !important;
            font-weight: 600;
        }}
        .stDownloadButton > button:hover, .stButton > button:hover {{
            background: #1e293b !important;
            color: #ffffff !important;
            border-color: #1e293b !important;
        }}
        .stDownloadButton button *, .stButton button * {{
            color: #ffffff !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


brand_name = "Excel Logics"
primary_color = "#0099CC"
secondary_color = "#B3CD36"
logo_upload = None

_inject_brand_css(primary_color, secondary_color)
display_brand_name = brand_name.strip() or "Excel Logics"
logo_bytes, logo_mime = _resolve_logo_bytes(logo_upload)
logo_uri = _logo_data_uri(logo_bytes, logo_mime)

header_col1, header_col2 = st.columns([1, 6])
with header_col1:
    if logo_bytes:
        st.image(logo_bytes, width=110)
    else:
        st.image(_placeholder_logo(), width=110)
with header_col2:
    st.title(f"{display_brand_name} - AI Excel Practice Bot")

roles = [
    "Data Analyst",
    "Sales Executive",
    "HR Associate",
    "Finance Associate",
    "Operations Associate",
    "Student (General)",
]

with st.form("student-form", clear_on_submit=False):
    c1, c2 = st.columns(2)
    with c1:
        student_name = st.text_input("Student Name *")
        phone = st.text_input("Phone Number *")
        is_working = st.selectbox("Are you currently working?", ["No", "Yes"])

        company_name = ""
        current_role = ""
        if is_working == "Yes":
            company_name = st.text_input("Company Name *")
            current_role = st.text_input("Current Role *")

    with c2:
        target_role = st.selectbox("Practice Dataset Role", roles)
        rows = st.slider("Rows to generate", min_value=200, max_value=10000, value=1000, step=100)
        st.text_input("Number of Questions", value=str(STANDARD_QUESTION_COUNT), disabled=True)

    student_question = st.text_area("Any simple question/request from student?")
    submit = st.form_submit_button("Generate Practice Excel")

if submit:
    errors = []
    if not student_name.strip():
        errors.append("Student Name is required.")
    if not _random_phone_ok(phone):
        errors.append("Enter a valid phone number (10-15 digits).")
    if is_working == "Yes" and (not company_name.strip() or not current_role.strip()):
        errors.append("Company Name and Current Role are required for working students.")

    if errors:
        for err in errors:
            st.error(err)
    else:
        unique_seed = _student_seed(student_name, phone, target_role)
        df = _generate_base_frame(target_role, rows, seed=unique_seed)
        used_signatures = _load_used_question_signatures()
        selected_questions, new_signatures = _build_question_bank(df, seed=unique_seed, used_signatures=used_signatures)
        if len(selected_questions) < STANDARD_QUESTION_COUNT:
            st.error(
                "Could not generate enough globally unique questions right now. "
                "Please try again with a different role or row count."
            )
            st.stop()

        _persist_used_question_signatures(new_signatures)

        profile_payload = {
            "SerialNumber": _next_queue_id(),
            "StudentName": student_name.strip(),
            "Phone": phone.strip(),
            "WorkingStatus": is_working,
            "CompanyName": company_name.strip() if is_working == "Yes" else "N/A",
            "CurrentRole": current_role.strip() if is_working == "Yes" else "N/A",
            "PracticeRole": target_role,
            "QuestionCount": str(STANDARD_QUESTION_COUNT),
            "StudentSeed": str(unique_seed),
            "StudentQuestion": student_question.strip() or "N/A",
            "GeneratedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        file_bytes = _build_workbook_bytes(profile_payload, df, selected_questions)
        safe_name = "".join(ch for ch in student_name if ch.isalnum())[:20] or "student"
        suffix = "".join(random.choices(string.ascii_lowercase + string.digits, k=4))
        filename = f"{safe_name}_{target_role.replace(' ', '_')}_{suffix}.xlsx"

        st.success("Practice workbook generated.")
        st.caption(f"Global unique question registry size: {len(used_signatures | new_signatures)}")
        st.download_button(
            "Download Practice Excel",
            data=file_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

st.divider()
st.subheader("Upload Solved File for Evaluation")
uploaded = st.file_uploader("Upload completed workbook (.xlsx)", type=["xlsx"])

if uploaded is not None:
    try:
        report, percentage, correct_count, incorrect_count, profile = _evaluate_workbook(uploaded.read())
        serial_number = _resolve_serial_number(profile)
        exam_name = "Excel Logics Skill Assessment"
        result_status = "Passed" if percentage > 70 else "Failed"
        cert_logo_html = f'<img class="cert-logo" src="{logo_uri}" alt="Excel Logics Logo" />' if logo_uri else ""

        left, right = st.columns([2, 3])
        with left:
            st.markdown(
                f"""
                <div class="certificate-card">
                    {cert_logo_html}
                    <p class="certificate-title">Practice Completion Certificate</p>
                    <p class="certificate-sub"><strong>Institute:</strong> {display_brand_name}</p>
                    <p class="certificate-sub"><strong>Exam Name:</strong> {exam_name}</p>
                    <p class="certificate-sub"><strong>Serial Number:</strong> {serial_number}</p>
                    <p class="certificate-sub"><strong>Name:</strong> {profile.get("StudentName", "N/A")}</p>
                    <p class="certificate-sub"><strong>Phone:</strong> {profile.get("Phone", "N/A")}</p>
                    <p class="certificate-sub"><strong>Working:</strong> {profile.get("WorkingStatus", "N/A")}</p>
                    <p class="certificate-sub"><strong>Company:</strong> {profile.get("CompanyName", "N/A")}</p>
                    <p class="certificate-sub"><strong>Current Role:</strong> {profile.get("CurrentRole", "N/A")}</p>
                    <p class="certificate-sub"><strong>Practice Role:</strong> {profile.get("PracticeRole", "N/A")}</p>
                    <p class="certificate-sub"><strong>Result:</strong> {result_status}</p>
                    <p class="certificate-score">{percentage}%</p>
                    <p class="certificate-sub">Correct Answers: <strong>{correct_count}</strong></p>
                    <p class="certificate-sub">Incorrect Answers: <strong>{incorrect_count}</strong></p>
                    <p class="certificate-sub"><strong>Generated At:</strong> {profile.get("GeneratedAt", "N/A")}</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.progress(min(max(percentage / 100.0, 0.0), 1.0))
            card_png = _build_scorecard_png(
                brand_name=display_brand_name,
                logo_bytes=logo_bytes,
                serial_number=serial_number,
                student_name=profile.get("StudentName", "Student"),
                practice_role=profile.get("PracticeRole", "N/A"),
                percentage=percentage,
                correct_count=correct_count,
                incorrect_count=incorrect_count,
                primary_color=primary_color,
                secondary_color=secondary_color,
                exam_name=exam_name,
                result_status=result_status,
            )
            st.download_button(
                "Download LinkedIn Score Card (PNG)",
                data=card_png,
                file_name=f"{profile.get('StudentName', 'student').replace(' ', '_')}_score_card.png",
                mime="image/png",
            )

        with right:
            chart_df = pd.DataFrame(
                {"Count": [int(correct_count), int(incorrect_count)]}, index=["Correct", "Incorrect"]
            )
            st.markdown("#### Score Breakdown")
            st.bar_chart(chart_df)

        st.markdown("#### Detailed Evaluation")
        display_cols = ["QuestionNo", "QuestionID", "ExpectedAnswer", "StudentAnswer", "Result"]
        st.dataframe(report[display_cols], use_container_width=True, hide_index=True)

        report_bytes = BytesIO()
        with pd.ExcelWriter(report_bytes, engine="openpyxl") as writer:
            pd.DataFrame(
                {
                    "Metric": ["Serial Number", "Exam Name", "Result", "Percentage", "Correct", "Incorrect", "Total"],
                    "Value": [serial_number, exam_name, result_status, percentage, correct_count, incorrect_count, correct_count + incorrect_count],
                }
            ).to_excel(writer, index=False, sheet_name="Summary")
            report.to_excel(writer, index=False, sheet_name="Evaluation")

        report_bytes.seek(0)
        st.download_button(
            "Download Evaluation Report",
            data=report_bytes.getvalue(),
            file_name="evaluation_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        st.error(f"Could not evaluate workbook: {exc}")
